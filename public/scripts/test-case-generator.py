#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例生成器 v1.0
基于中煤科工项目测试用例模板
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import sys

class TestCaseGenerator:
    """测试用例生成器"""
    
    def __init__(self, customer_name, module, test_date=None):
        self.customer_name = customer_name
        self.module = module
        self.test_date = test_date or datetime.now().strftime("%Y%m%d")
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "测试用例"
        self._setup_styles()
    
    def _setup_styles(self):
        """设置样式"""
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _create_header(self):
        """创建表头"""
        headers = [
            '流程代码', '步骤号', '步骤名称', '系统路径（详细）',
            '测试数据', '期望结果', '实际结果', '测试参与者', '测试评价', '测试完成日期'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # 设置列宽
        column_widths = [15, 8, 25, 30, 30, 30, 20, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 设置行高
        self.ws.row_dimensions[3].height = 30
    
    def generate(self, test_cases):
        """生成测试用例"""
        # 标题
        self.ws.merge_cells('A1:J1')
        title_cell = self.ws.cell(row=1, column=1, value=f"{self.customer_name}{self.module}测试用例")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[1].height = 30
        
        # 信息行
        self.ws.merge_cells('A2:B2')
        self.ws.cell(row=2, column=1, value="流程代码")
        self.ws.merge_cells('C2:D2')
        self.ws.cell(row=2, column=3, value="")
        self.ws.cell(row=2, column=5, value="流程所属部门")
        self.ws.merge_cells('F2:G2')
        self.ws.cell(row=2, column=6, value="")
        self.ws.cell(row=2, column=8, value="计划测试日期")
        self.ws.merge_cells('I2:J2')
        self.ws.cell(row=2, column=9, value="")
        
        # 创建表头
        self._create_header()
        
        # 填充测试用例
        row = 4
        for test_case in test_cases:
            self.ws.cell(row=row, column=1, value=test_case.get('process_code', ''))
            self.ws.cell(row=row, column=2, value=test_case.get('step_no', ''))
            self.ws.cell(row=row, column=3, value=test_case.get('step_name', ''))
            self.ws.cell(row=row, column=4, value=test_case.get('system_path', ''))
            self.ws.cell(row=row, column=5, value=test_case.get('test_data', ''))
            self.ws.cell(row=row, column=6, value=test_case.get('expected_result', ''))
            self.ws.cell(row=row, column=7, value=test_case.get('actual_result', ''))
            self.ws.cell(row=row, column=8, value=test_case.get('tester', ''))
            self.ws.cell(row=row, column=9, value=test_case.get('evaluation', ''))
            self.ws.cell(row=row, column=10, value=test_case.get('complete_date', ''))
            
            # 设置边框和对齐
            for col in range(1, 11):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            row += 1
        
        # 保存文件
        filename = f"{self.customer_name}_{self.module}_测试用例_{self.test_date}.xlsx"
        self.wb.save(filename)
        return filename


def get_module_test_cases(module):
    """获取模块测试用例模板"""
    
    templates = {
        "出纳": [
            {
                'process_code': 'BTF_13',
                'step_no': '1.1',
                'step_name': '付款单校验资金计划，提交审核后推送司库',
                'system_path': '财务云-出纳-付款处理',
                'test_data': '选择结算方式为"电汇"的付款排程下推的付款单',
                'expected_result': '在付款单可以查看资金计划，超过资金进行提示。资金计划内的单据状态变化：已提交-已审核-付款处理中',
                'actual_result': '',
                'tester': '',
                'evaluation': '',
                'complete_date': ''
            },
            {
                'process_code': 'BTF_13',
                'step_no': '1.2',
                'step_name': '付款单审批流程测试',
                'system_path': '财务云-出纳-付款处理',
                'test_data': '提交付款单，触发审批流程',
                'expected_result': '审批流程按配置正确流转，审批记录完整',
                'actual_result': '',
                'tester': '',
                'evaluation': '',
                'complete_date': ''
            },
        ],
        "应收": [
            {
                'process_code': 'STC_07',
                'step_no': '1.1',
                'step_name': '正常开票流程测试',
                'system_path': '财务云-应收管理-发票管理',
                'test_data': '创建销售发票，关联销售出库单',
                'expected_result': '发票创建成功，关联关系正确，金额一致',
                'actual_result': '',
                'tester': '',
                'evaluation': '',
                'complete_date': ''
            },
        ],
        "应付": [
            {
                'process_code': 'PTM_14',
                'step_no': '1.1',
                'step_name': '应付挂账流程测试',
                'system_path': '财务云-应付管理-应付单',
                'test_data': '创建应付单，关联采购入库单',
                'expected_result': '应付单创建成功，关联关系正确，金额一致',
                'actual_result': '',
                'tester': '',
                'evaluation': '',
                'complete_date': ''
            },
        ],
        "总账": [
            {
                'process_code': 'BTF_38',
                'step_no': '1.1',
                'step_name': '凭证录入测试',
                'system_path': '财务云-总账-凭证处理',
                'test_data': '录入测试凭证，验证借贷平衡',
                'expected_result': '凭证保存成功，借贷平衡校验通过',
                'actual_result': '',
                'tester': '',
                'evaluation': '',
                'complete_date': ''
            },
        ],
    }
    
    return templates.get(module, templates["总账"])


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='测试用例生成器 v1.0')
    parser.add_argument('--customer', required=True, help='客户名称')
    parser.add_argument('--module', required=True,
                       choices=['出纳', '应收', '应付', '总账', '费用', '预算', '成本', '资产'],
                       help='测试模块')
    parser.add_argument('--date', help='测试日期 (YYYYMMDD)')
    
    args = parser.parse_args()
    
    generator = TestCaseGenerator(
        customer_name=args.customer,
        module=args.module,
        test_date=args.date
    )
    
    # 获取测试用例模板
    test_cases = get_module_test_cases(args.module)
    
    filename = generator.generate(test_cases)
    print(f"测试用例已生成: {filename}")


if __name__ == "__main__":
    main()
